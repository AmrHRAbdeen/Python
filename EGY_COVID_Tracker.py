#############################################################
# Developing a script to GET my Country COVID_19 statistics 
#############################################################
import requests
import pandas
import datetime
from openpyxl import Workbook
from openpyxl import load_workbook
import matplotlib.pyplot as plt 

# Get current date
currentDayDate = datetime.datetime.now()
# Check if the Database already existed 
workBook = load_workbook('EGY_COVID_19_Database.xlsx')   
print(workBook.sheetnames)
if ('EGY_COVID_19_Database' == workBook.sheetnames[0] ):
    print("File already Existed ...")
    workSheet=workBook.active
else:
    print("WB not here , Creating Database ...")
    # Store Date in Excel sheet
    # 1- Create Workbook
    workBook = Workbook()
    # 2- Change Worksheet name
    # Workbook contains multiple worksheets we need to get the active one using active property
    workSheet=workBook.active
    workSheet.title="EGY_COVID_19_Database"

# x axis values 
DateDaily =[] 
# corresponding y axis values 
totalInfectedCases =[]
# Date Counter
dateCounter=0
#Determine Worksheet Cell for date
dateRow=1
dateCol=1
dateCell=workSheet.cell(row=dateRow,column=dateCol)
dateCell.value ='Date'
totalrow=1
totalcol=2
totalCell = workSheet.cell(row=totalrow,column=totalcol)
totalCell.value ='Total Cases'
globalEgyIndex=0
URL ="https://www.worldometers.info/coronavirus/"
print("Getting Egypt COVID_19 statistics ...")
pageResponse = requests.get(URL)
pageContent = pageResponse.text
#print(pageContent)
dataFrames=  pandas.read_html(pageContent)
dataFrameOfInterest=dataFrames[0]
#print(dataFrames)
#print(dataFrames[0].loc[0][0])
NumofRows=len(dataFrameOfInterest.index)
NumOfCol=len(dataFrameOfInterest.columns)
print("Num of cols = {}".format(NumOfCol))
for cnt in range(0,NumOfCol):
    for inCnt in range(0,NumofRows):
        reeee= dataFrames[0].loc[inCnt][cnt]
        if( reeee == 'Egypt'):
            globalEgyIndex = inCnt
            print("Egypt index is ={}".format(inCnt))
            print("Egypt No of total cases of COVID_19 is :{}".format(dataFrames[0].loc[inCnt][cnt+1]))
            #3- Store our Data [Date(A) , totalCases(B)]
            while (dateCell.value != None):
                if (dateCell.value == 'Date'):
                    dateRow+=1
                    dateCell=workSheet.cell(row=dateRow,column=dateCol)
                    continue
                dateCell=workSheet.cell(row=dateRow,column=dateCol)
                dayDate=str(dateCell.value)
                getDayDate=dayDate[8:10]
                #print(getDayDate)
                dateCounter+=1
                print("dateCounter={}".format(dateCounter))
                DateDaily.append(dateCounter)
                dateRow+=1
                dateCell=workSheet.cell(row=dateRow,column=dateCol) 
            dateCell.value=str(currentDayDate)
            dateCell=workSheet.cell(row=dateRow,column=dateCol)
            dayDate=str(dateCell.value)
            getDayDate=dayDate[8:10]
            DateDaily.append(getDayDate)
            print("Date list len = {}".format(len(DateDaily)))
            while (totalCell.value != None):
                if (totalCell.value == 'Total Cases'):
                    totalrow+=1
                    totalCell=workSheet.cell(row=totalrow,column=totalcol)
                    continue
                else:    
                    totalInfectedCases.append(totalCell.value)
                    print(totalCell.value)
                    totalrow+=1
                    totalCell=workSheet.cell(row=totalrow,column=totalcol)
            totalCell.value =dataFrames[0].loc[inCnt][cnt+1]
            totalInfectedCases.append(dataFrames[0].loc[inCnt][cnt+1])
            print("totalInfectedCases list len = {}".format(len(totalInfectedCases)))
            print("Egypt No of New cases of COVID_19 is : {}".format(dataFrames[0].loc[inCnt][cnt+2])) 
        
            
################## Plotting the exp. function expectations 
#  y =a^x : duplicate => y= 2^x , x range= [0, +inf. ]  => y =2 ** x###########
expectedDays=[]
expectedCases=[]
# Apply exp. function only if total Cases > 1000 
# Duration 15 Days
if dataFrames[0].loc[globalEgyIndex][2] > 800:
    for cnt in range(0,15):
        expectedDays.append(cnt)
        expectedVal = 2 ** cnt
        expectedCases.append(expectedVal)
        print(expectedDays[cnt])
        print(expectedCases[cnt])
#plt.plot(expectedDays, expectedCases , 'r')
plt.plot(DateDaily, totalInfectedCases , 'b') 
    
# plotting the points  

# naming the x axis 
plt.xlabel('Days Counter') 
# naming the y axis 
plt.ylabel('Total Infected Cases') 
  
# giving a title to my graph 
plt.title('Egypt COVID_19 Graph') 

plt.grid(True)  
# function to show the plot 
plt.show() 
#4- Save File/workBook
workBook.save(filename='EGY_COVID_19_Database.xlsx')




